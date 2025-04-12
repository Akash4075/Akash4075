import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# Excel file name
excel_file = "Doctor_Prescriptions.xlsx"

# Create Excel file if not exists
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Patient Name", "Phone Number", "Location", "Prescription"])
    wb.save(excel_file)

# Save to Excel function
def save_prescription():
    name = entry_name.get().strip()
    number = entry_number.get().strip()
    location = entry_location.get().strip()
    prescription = text_prescription.get("1.0", tk.END).strip()
    
    if not name or not number or not location or not prescription:
        messagebox.showwarning("Missing Info", "Please fill in all fields.")
        return

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([date, time, name, number, location, prescription])
    wb.save(excel_file)

    messagebox.showinfo("Saved", "Prescription saved successfully.")
    entry_name.delete(0, tk.END)
    entry_number.delete(0, tk.END)
    entry_location.delete(0, tk.END)
    text_prescription.delete("1.0", tk.END)

# Tkinter GUI
root = tk.Tk()
root.title("Doctor Prescription Entry")
root.geometry("500x600")
root.configure(bg="#E6F2FF")  # Light blue background

tk.Label(root, text="Doctor Prescription Entry", font=("Helvetica", 18, "bold"), bg="#E6F2FF", fg="black").pack(pady=10)

tk.Label(root, text="Patient Name:", bg="#E6F2FF").pack(anchor="w", padx=20)
entry_name = tk.Entry(root, width=50)
entry_name.pack(padx=20, pady=5)

tk.Label(root, text="Phone Number:", bg="#E6F2FF").pack(anchor="w", padx=20)
entry_number = tk.Entry(root, width=50)
entry_number.pack(padx=20, pady=5)

tk.Label(root, text="Location:", bg="#E6F2FF").pack(anchor="w", padx=20)
entry_location = tk.Entry(root, width=50)
entry_location.pack(padx=20, pady=5)

tk.Label(root, text="Prescription:", bg="#E6F2FF").pack(anchor="w", padx=20)
text_prescription = tk.Text(root, width=58, height=10)
text_prescription.pack(padx=20, pady=5)

tk.Button(root, text="Save Prescription", command=save_prescription, bg="green", fg="white", font=("Helvetica", 12, "bold")).pack(pady=20)

root.mainloop()
